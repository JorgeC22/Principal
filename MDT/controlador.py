from openpyxl import load_workbook
from datetime import datetime
from conexion import *
import os
import sys

def cargar(file, DB):
    now = datetime.now()
    print(f"{now.strftime('%Y-%m-%d %H:%M:%S')} Inicio de lectura de archivo ")
    wb = load_workbook('./archivos/'+file.filename)    
    sheet = wb.active
    #inicio = 4
    fin = sheet.max_row
    data = [] # lista de tuplas
    now = datetime.now()
    listaFilas = list(sheet.rows)
    print(f"{now.strftime('%Y-%m-%d %H:%M:%S')} se encontraron {fin} registros, se inicia carga en memoria de {len(listaFilas)-4} filas\n")

    j=1
    for row in listaFilas[4:fin]:
        print(f"\rLeyendo fila {j} de {fin-4}", end = '', flush=True)
        sys.stdout.flush()
        datosFila = []
        #print(row)
        #print(f'fila {j}')

        for valor in row:
            datosFila.append(valor.value) #se llena el arreglo correspondiente a la fila
            
        #try:
            #datosFila[2] = datetime.strptime(str(datosFila[2]), "%d/%m/%y").strftime("%Y-%m-%d")
            #datosFila[3] = datetime.strptime(str(datosFila[3]), "%d/%m/%y").strftime("%Y-%m-%d")
        #except BaseException as e:
            #print(f"error al transformar fecha {e}")
            #sys.exit(0)

        data.append(tuple(datosFila))
        j+=1

    now = datetime.now()
    print(f"\n{now.strftime('%Y-%m-%d %H:%M:%S')} terminada la carga en memoria, se inicia carga en BD")

    if os.path.exists('./archivos/'+file.filename):
        os.remove('./archivos/'+file.filename)

    totalInserciones = len(data)
    
    u= 1
    for i in data:
        print(f"\rInsertando registro {u} de {totalInserciones}", end = '', flush=True)
        sys.stdout.flush()
        #existe = DB.existeRegistro(i)
        #print(f'existe = {existe}')
        #DB.insertaRegistro(tuplaDatos = i)
        u+=1
        if DB.existeRegistro(i)['cuenta'] == 0:
            DB.insertaRegistro(tuplaDatos = i)
            u+=1
    now = datetime.now()
    print(f"\n{now.strftime('%Y-%m-%d %H:%M:%S')} terminada la carga en BD")

def main(archivo):
    file = archivo
    DB = dbMDT()
    cargar(file, DB)